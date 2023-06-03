# This Python file uses the following encoding: utf-8
import sys, os, glob, re, datetime, time, shutil

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.styles import numbers

from PySide6.QtCore import Qt, Signal, QThread
from PySide6.QtPdf import QPdfDocument
from PySide6.QtGui import QTextCursor, QMouseEvent, QKeyEvent
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QDialog, QDialogButtonBox, QMessageBox, QLabel,
    QVBoxLayout, QHBoxLayout,
    QListWidget,
    QListWidgetItem, QTableWidgetItem,
    QFileDialog,
    QSizeGrip
)

# Important:
# You need to run the following command to generate the ui_form.py file
#     pyside6-uic form.ui -o ui_form.py, or
#     pyside2-uic form.ui -o ui_form.py
from ui_form import Ui_MainWindow
from ui_dialog import Ui_Dialog, Success


class FolderScanner(QThread) :
    process = Signal(dict)
    def __init__(self, location:str) :
        super().__init__()
        self.location = location
    
    def run(self) :
        if self.location != '' :
            while True :
                files = glob.glob(self.location + '/*.pdf')
                container = {os.path.split(x)[1] : x for x in files}
                self.process.emit(container)

    def stop(self) :
        self.finished.emit()
        self.terminate()


class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.setWindowFlag(Qt.WindowType.FramelessWindowHint, True)
        grip = QSizeGrip(self.ui.widgetEditor)
        self.ui.layoutEditor.addWidget(grip, alignment=(Qt.AlignmentFlag.AlignBottom | Qt.AlignmentFlag.AlignRight))

        # Variables
        self.bulan = {
            1:'Januari', 2: 'Februari', 3: 'Maret', 4: 'April', 5: 'May', 6: 'Juni',
            7: 'Juli', 8: 'Agustus', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }
        self.currentFolder = ''
        self.pdfContainerProses = {}
        self.pdfContainerSelesai = {}
        self.excelLocation = ('', '')
        self.sekarang = f'{datetime.date.today().day} {self.bulan[datetime.date.today().month]}'

        # Title Bar
        self.ui.tbExit.clicked.connect(self.close)
        self.ui.tbMaximize.clicked.connect(lambda : self.showNormal() if self.isMaximized() else self.showMaximized())
        self.ui.tbMinimize.clicked.connect(self.showMinimized)
        self.ui.widgetTitleBar.mousePressEvent = self.windowClicked
        self.ui.widgetTitleBar.mouseMoveEvent = self.moveWindow
        self.currentPositionX = 0
        self.currentPositionY = 0

        # Button
        self.ui.widgetShortcut.setVisible(False)
        self.ui.tombolShortcut.clicked.connect(lambda : self.buttonVisibility(False))
        self.ui.tombolKembali.clicked.connect(lambda : self.buttonVisibility(True))
        self.ui.tombolFolder.clicked.connect(self.pilihFolder)

        # About
        self.ui.labelLink.setTextFormat(Qt.TextFormat.RichText)
        self.ui.labelLink.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        self.ui.labelLink.setOpenExternalLinks(True)

        # List Widget
        self.ui.listWidgetProses.currentItemChanged.connect(lambda pdf : self.showPdfContent(pdf, self.pdfContainerProses))
        self.ui.listWidgetSelesai.currentItemChanged.connect(lambda pdf : self.showPdfContent(pdf, self.pdfContainerSelesai))

        # Text Editor
        self.ui.textEdit.setReadOnly(True)
        self.ui.textEdit.keyPressEvent = self.shortcutKey

        # Tools
        self.ui.tbTambahRow.clicked.connect(self.addRow)
        self.ui.tbHapusRow.clicked.connect(self.delRow)
        self.ui.tbOpen.clicked.connect(self.openExcel)
        self.ui.tbSave.clicked.connect(self.saveExcel)
        self.ui.tbClear.clicked.connect(self.clearTable)
        self.ui.leCell.setPlaceholderText('A1')
        self.ui.leFileName.setPlaceholderText(self.sekarang)

        # Tabel
        self.ui.tableWidget.setCurrentCell(0, 0)

        # Tab Widget
        self.ui.tabWidget.currentChanged.connect(self.tabChanged)



##########
# Method #
##########

    # Title Bar
    def windowClicked(self, event: QMouseEvent) -> None:
        if event.type() == event.Type.MouseButtonDblClick :
            if self.isMaximized() : self.showNormal()
            else : self.showMaximized()

        self.currentPositionX = event.position().x() + self.ui.widgetTombol.size().width()
        self.currentPositionY = event.position().y()

    def moveWindow(self, event: QMouseEvent) -> None :
        self.move(event.globalPosition().x() - self.currentPositionX, event.globalPosition().y() - self.currentPositionY)



    # Button
    def buttonVisibility(self, state:bool) :
        self.ui.widgetShortcut.setVisible(not state)
        self.ui.tombolFolder.setVisible(state)
        self.ui.tombolShortcut.setVisible(state)
    
    def pilihFolder(self) :
        currentFolder = QFileDialog.getExistingDirectory()
        if currentFolder != self.currentFolder and currentFolder != '' :
            self.pdfContainerProses.clear()
            self.ui.listWidgetProses.clear()
            self.currentFolder = currentFolder.replace('/', '\\')
            if hasattr(self, 'scanner') :
                self.scannerProcess.stop()
                self.scannerSelesai.stop()
            if not os.path.exists(self.currentFolder + '\\' + self.sekarang) :
                os.mkdir(self.currentFolder + '\\' + self.sekarang)
            self.scannerProcess = FolderScanner(self.currentFolder)
            self.scannerSelesai = FolderScanner(self.currentFolder + "\\" + self.sekarang)
            self.scannerProcess.process.connect(lambda pdf: self.addPdfTo(pdf, self.pdfContainerProses, self.ui.listWidgetProses))
            self.scannerSelesai.process.connect(lambda pdf: self.addPdfTo(pdf, self.pdfContainerSelesai, self.ui.listWidgetSelesai))
            self.scannerProcess.start()
            self.scannerSelesai.start()


    # List PDF
    def addPdfTo(self, pdf:dict, container:dict, widget:QListWidget) :
        listPdf = widget
        if container :
            try :
                for file in container :
                    if file not in pdf :
                        for i in range(listPdf.count()) :
                            if listPdf.item(i) == None :
                                pass
                            elif listPdf.item(i).text() == file :
                                listPdf.takeItem(i)
                                container.pop(file)
                for file in pdf :
                    if file not in container :
                        container[file] = pdf[file]
                        listPdf.addItem(QListWidgetItem(file))
            except RuntimeError :
                print("Dictionary Berubah Saat Melakukan Looping")
        else :
            for file in pdf :
                container[file] = pdf[file]
                listPdf.addItem(QListWidgetItem(file))

    def showPdfContent(self, currentList:QListWidgetItem, container:dict) :
        self.ui.textEdit.clear()
        if hasattr(currentList, 'text') :
            currentPdf = container[currentList.text()]
            pdf = QPdfDocument(self)
            pdf.load(currentPdf)
            for i in range(pdf.pageCount()) :
                self.ui.textEdit.insertPlainText(pdf.getAllText(i).text())
            pdf.close()
            pdf.deleteLater()


    # Tab Widget
    def tabChanged(self) :
        self.ui.textEdit.clear()
        self.ui.listWidgetProses
        self.ui.listWidgetSelesai


    # Text Editor
    def getSelection(self, col:int) :
        if self.ui.textEdit.textCursor().hasSelection() :
            text = self.ui.textEdit.textCursor().selection().toPlainText()
            tableWidget = self.ui.tableWidget
            tableWidget.setItem(tableWidget.currentRow(), col, QTableWidgetItem(text))
            self.ui.tableWidget.resizeRowsToContents()

    # Field
    def fieldValidator(self) :
        cell = self.ui.leCell
        name = self.ui.leFileName

        dialog = QMessageBox(self)
        dialog.setWindowTitle(u'Peringatan!')
        dialog.setIcon(QMessageBox.Icon.Warning)
        dialog.setStandardButtons(QMessageBox.StandardButton.Ok)

        if not re.match('^[a-zA-Z]{1,3}[1-9]{1}[0-9]{0,8}$', cell.text()) and cell.text() != '' :
            cell.setText('')
            dialog.setText('Format Cell tidak Valid!')
            dialog.exec()
            return False
        if re.fullmatch('[\\/:*?"<>|]', name.text()) :
            dialog.setText('Nama berisi karakter tidak Valid!')
            dialog.exec()
            return False
        return True


    # Table
    def addRow(self) :
        tabel = self.ui.tableWidget
        tabel.setRowCount(tabel.rowCount() + 1)

    def delRow(self) :
        tabel = self.ui.tableWidget
        if tabel.rowCount() > 1 :
            if tabel.currentRow() == -1 :
                tabel.setRowCount(tabel.rowCount() - 1)
            else :
                tabel.removeRow(tabel.currentRow())
                tabel.setCurrentCell(tabel.currentRow(), 0)
        else :
            tabel.clearContents()
            tabel.setCurrentCell(0, 0)
    
    def nextRow(self) :
        tabel = self.ui.tableWidget
        if tabel.currentRow() < tabel.rowCount() - 1 :
            tabel.setCurrentCell(tabel.currentRow() + 1, 0)

    def previousRow(self) :
        tabel = self.ui.tableWidget
        if tabel.currentRow() > 0 :
            tabel.setCurrentCell(tabel.currentRow() - 1, 0)

    def clearTable(self) :
        for i in range(self.ui.tableWidget.rowCount()) :
            self.ui.tableWidget.removeRow(i)
        self.ui.tableWidget.removeRow(0)
        self.ui.tableWidget.setRowCount(1)


    # Excel
    def openExcel(self) :
        excelLocation = (QFileDialog.getOpenFileName(self, caption='Pilih File Excel', filter='*.xlsx')[0]).replace('/', '\\')
        if (excelLocation != '') and (excelLocation != self.excelLocation) :
            self.excelLocation = excelLocation
            self.ui.leFileName.setPlaceholderText((os.path.split(self.excelLocation)[1]).rstrip('.xlsx'))
            print(self.excelLocation)

    def saveExcel(self) :
        if self.fieldValidator() :
            if self.dialogSimpan() :
                if self.excelLocation == ('', '') :
                    excel = Workbook()
                    self.tabelToExcel(excel)
                    if self.ui.leFileName.text() == '' :
                        if self.currentFolder :
                            if not os.path.exists(self.currentFolder + '\\Data Excel') :
                                os.mkdir(self.currentFolder + '\\Data Excel')
                            location = self.currentFolder + '\\Data Excel\\' + self.ui.leFileName.placeholderText().rstrip('.xlsx')
                            path = self.currentFolder
                        else :
                            location = self.ui.leFileName.placeholderText().rstrip('.xlsx')
                            path = os.path.split(__file__)[0]
                        print(location + '.xlsx')
                        print(path + '\\Data Excel')
                        excel.save(location + '.xlsx')
                        dialog = Success(path + '\\Data Excel')
                        dialog.show()
                        dialog.exec()
                        dialog.deleteLater()
                        excel.close()
                    else :
                        if self.currentFolder :
                            if not os.path.exists(self.currentFolder + '\\Data Excel') :
                                os.mkdir(self.currentFolder + '\\Data Excel')
                            location = self.currentFolder + '\\Data Excel\\'  + self.ui.leFileName.text().rstrip('.xlsx')
                            path = self.currentFolder
                        else :
                            location = self.ui.leFileName.text().rstrip('.xlsx')
                            path = os.path.split(__file__)[0]
                        print(location + '.xlsx')
                        print(path + '\\Data Excel')
                        excel.save(location + '.xlsx')
                        dialog = Success(path + '\\Data Excel')
                        dialog.show()
                        dialog.exec()
                        dialog.deleteLater()
                        excel.close()
                else :
                    if not os.path.isfile(self.excelLocation) :
                        excel = Workbook()
                    else :
                        excel = load_workbook(self.excelLocation)
                    excelPath = os.path.split(self.excelLocation)[0].replace('/', '\\') + '\\'
                    excelName = os.path.split(self.excelLocation)[0].rstrip('.xlsx')
                    fieldName = self.ui.leFileName.text().rstrip('.xlsx')
                    print("Path : " + excelPath)
                    print("Name " + excelName)
                    print(fieldName)
                    save = self.excelLocation
                    if fieldName != '' and fieldName != excelName:
                        save = excelPath + fieldName + '.xlsx'
                    location = os.path.split(save)[0]
                    try :
                        self.tabelToExcel(excel)
                        print(location)
                        print(save)
                        excel.save(save)
                        dialog = Success(location)
                        dialog.show()
                        dialog.exec()
                        dialog.deleteLater()
                        excel.close()
                    except PermissionError :
                        dialog = QMessageBox(self)
                        dialog.setWindowTitle(u'Peringatan!')
                        dialog.setText('Harap Tutup Dulu Microsoft Excel nya!')
                        dialog.setIcon(QMessageBox.Icon.Warning)
                        dialog.setStandardButtons(QMessageBox.StandardButton.Ok)
                        dialog.exec()
            else :
                pass
        else :
            print('Validation Failed!')

    def tabelToExcel(self, excel:Workbook) :
        sheet = excel.active
        table = self.ui.tableWidget
        rowCount = table.rowCount()
        colCount = table.columnCount()
        cellPH = self.ui.leCell.placeholderText()
        cellTx = self.ui.leCell.text().upper()
        startCell = coordinate_from_string(cellPH)

        if cellTx != '' :
            startCell = coordinate_from_string(cellTx)

        colNum = column_index_from_string(startCell[0])

        for row in range(rowCount) :
            for col in range(colCount) :
                cell = get_column_letter(colNum + col) + str(startCell[1] + row)
                if table.item(row, col) != None :
                    if table.item(row, col).text().isnumeric() :
                        sheet[cell].number_format = numbers.FORMAT_NUMBER
                        sheet[cell] = int(table.item(row, col).text())
                    else :
                        sheet[cell] = table.item(row, col).text()


    # Shortcut
    def shortcutKey(self, event:QKeyEvent) :
        if event.key() == Qt.Key.Key_A :
            self.addRow()
        if event.key() == Qt.Key.Key_D :
            self.delRow()
        if event.key() == Qt.Key.Key_W :
            self.previousRow()
        if event.key() == Qt.Key.Key_S :
            self.nextRow()
        if event.key() == Qt.Key.Key_1 :
            self.getSelection(0)
        if event.key() == Qt.Key.Key_2 :
            self.getSelection(1)
        if event.key() == Qt.Key.Key_3 :
            self.getSelection(2)
        if event.key() == Qt.Key.Key_Space :
            table = self.ui.tableWidget
            listWidget = self.ui.listWidgetProses
            currentTab = self.ui.tabWidget.currentIndex()
            if currentTab == 1 :
                listWidget = self.ui.listWidgetSelesai
            if not listWidget.item(listWidget.currentRow()) == None :
                key = listWidget.item(listWidget.currentRow()).text()
                if currentTab == 0 :
                    self.moveTo(self.pdfContainerProses, self.pdfContainerSelesai, key, 0)
                if currentTab == 1 :
                    self.moveTo(self.pdfContainerSelesai, self.pdfContainerProses, key, 1)
                listWidget.takeItem(listWidget.currentRow())

                if table.currentRow() == table.rowCount() - 1 :
                    self.addRow()
                table.setCurrentCell(table.currentRow() + 1, 0)

        if self.focusWidget().objectName() == 'textEdit' :
            if event.key() == Qt.Key.Key_Shift :
                self.ui.textEdit.moveCursor(QTextCursor.MoveOperation.StartOfBlock, QTextCursor.MoveMode.MoveAnchor)

    # Memindahkan File
    def moveTo(self, source:dict, destination:dict, key:str, flag:int) :
        now = datetime.datetime.now()
        pattern = '\[[0-9]{1,2}_[0-9]{1,2}_[0-9]{1,2}\]'
        if re.search(pattern, key) :
            new = key.rstrip('.pdf').replace(re.search(pattern, key).group(), f'[{now.hour}_{now.minute}_{now.second}]') + '.pdf'
            old = key
        else :
            new = key.rstrip('.pdf') + f' [{now.hour}_{now.minute}_{now.second}]' + '.pdf'
            old = key
        folder = self.currentFolder + '\\'
        subFolder = folder + self.sekarang + '\\'
        try :
            if flag == 0 :
                shutil.move(folder + old, subFolder + new)
            elif flag == 1 :
                shutil.move(subFolder + old, folder + old)
        except TypeError :
            print("Gagal Memindahkan File")

    # Dialog Save
    def dialogSimpan(self) :
        nthCell = self.ui.leCell.placeholderText()
        excelName = self.ui.leFileName.placeholderText()
        if self.ui.leCell.text() != '' :
            nthCell = self.ui.leCell.text()
        if self.ui.leFileName.text() != '' :
            excelName = self.ui.leFileName.text()
        dialog = Ui_Dialog(nthCell, excelName)
        status = dialog.exec()
        return status
    



if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = MainWindow()
    widget.show()
    sys.exit(app.exec())
